"""内容清洗模块——对不同类型文件执行针对性清洗，去除噪声信息。

核心原则：不拿原始垃圾数据去向量化浪费 Token。
每种文件类型有独立的清洗策略，返回干净的可读文本。
"""

from __future__ import annotations

import csv
import io
import re
from typing import Any

# 可选依赖——首次使用时按需导入，缺失不阻塞模块加载


class CleanResult:
    """清洗结果。"""

    __slots__ = ("text", "warnings")

    def __init__(self, text: str, warnings: list[str] | None = None) -> None:
        self.text = text
        self.warnings = warnings or []


def clean_content(file_bytes: bytes, file_type: str) -> CleanResult:
    """根据文件类型分发到对应清洗策略。

    Args:
        file_bytes: 原始文件字节。
        file_type: 文件类型标识（pdf / docx / xlsx / csv / html / md / txt 或代码扩展名）。

    Returns:
        CleanResult，text 为清洗后可读字符串。
    """
    _normalized = file_type.lower().lstrip(".")
    if _normalized == "pdf":
        return _clean_pdf(file_bytes)
    if _normalized == "docx":
        return _clean_docx(file_bytes)
    if _normalized in ("xlsx", "xls"):
        return _clean_xlsx(file_bytes)
    if _normalized == "csv":
        return _clean_csv(file_bytes)
    if _normalized in ("html", "htm"):
        return _clean_html(file_bytes)
    if _normalized == "md":
        return _clean_markdown(file_bytes)
    if _normalized in {"txt", "log", "json", "yaml", "yml", "toml", "cfg", "ini", "conf", "env"}:
        return _clean_text(file_bytes)
    # 代码文件——保留原样（高信号），仅去连续多余空行
    if _normalized in _CODE_EXTENSIONS:
        return _clean_code(file_bytes)
    # 图片/二进制——首版跳过
    if _normalized in {"png", "jpg", "jpeg", "gif", "bmp", "svg", "ico", "webp", "mp4", "mp3", "zip", "tar", "gz", "bin", "exe", "dll"}:
        return CleanResult("", ["不支持的文件类型——图片和二进制文件暂不处理"])
    # 未知类型回退纯文本清洗
    return _clean_text(file_bytes)


# ── 各类型清洗实现 ──────────────────────────────────────────────────────────────


def _clean_pdf(data: bytes) -> CleanResult:
    """pymupdf 提取 PDF 文本并清洗：去页眉页脚、连字符断词、页码、空页。"""
    import fitz  # pymupdf

    warnings: list[str] = []
    doc = fitz.open(stream=data, filetype="pdf")
    pages: list[str] = []
    for _page_num, page in enumerate(doc):
        text = page.get_text("text")
        if not text.strip():
            continue
        # 去连字符断词（行末 hyphen + 换行）
        text = re.sub(r"(\w+)-\n(\w+)", r"\1\2", text)
        # 去纯数字页码行（独立成行且为数字）
        text = re.sub(r"^\s*\d{1,4}\s*$", "", text, flags=re.MULTILINE)
        # 合并空白行
        text = re.sub(r"\n{3,}", "\n\n", text)
        pages.append(text.strip())
    doc.close()
    if not pages:
        return CleanResult("", ["PDF 无可提取文本——可能为扫描件"])
    return CleanResult("\n\n".join(pages), warnings)


def _clean_docx(data: bytes) -> CleanResult:
    """python-docx 提取 DOCX 段落，保留标题层级标记。"""
    from docx import Document  # python-docx

    doc = Document(io.BytesIO(data))
    parts: list[str] = []
    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        # 标题前缀标记以保留结构
        if para.style.name.startswith("Heading"):
            level = para.style.name.split()[-1]
            if level.isdigit():
                text = f"{'#' * int(level)} {text}"
        parts.append(text)
    if not parts:
        return CleanResult("", ["DOCX 无可提取文本"])
    return CleanResult("\n\n".join(parts))


def _clean_xlsx(data: bytes) -> CleanResult:
    """openpyxl 读取 Excel，识别表头行后每行转为'列名: 值'结构化文本。"""
    from openpyxl import load_workbook  # openpyxl

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    all_rows_text: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        # 第一行作为表头
        header: list[str] = []
        try:
            first_row: tuple[Any, ...] = next(rows_iter)  # type: ignore[assignment]
            header = [str(h).strip() if h is not None else "" for h in first_row]
        except StopIteration:
            continue
        sheet_parts: list[str] = [f"[工作表: {sheet_name}]"]
        for row in rows_iter:
            row_vals: tuple[Any, ...] = row  # type: ignore[assignment]
            if all(v is None for v in row_vals):
                continue
            fields: list[str] = []
            for idx, val in enumerate(row_vals):
                col_name = header[idx] if idx < len(header) and header[idx] else f"列{idx}"
                val_str = str(val).strip() if val is not None else "(空)"
                fields.append(f"{col_name}: {val_str}")
            sheet_parts.append(" | ".join(fields))
        all_rows_text.extend(sheet_parts)
    wb.close()
    if not all_rows_text:
        return CleanResult("", ["Excel 无可读取数据"])
    return CleanResult("\n".join(all_rows_text))


def _clean_csv(data: bytes) -> CleanResult:
    """CSV ——同 Excel 策略，识别表头后结构化转换。"""
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        return CleanResult("", ["CSV 为空"])
    header = [h.strip() for h in header]
    rows: list[str] = []
    for row in reader:
        if all(not cell.strip() for cell in row):
            continue
        fields = []
        for idx, val in enumerate(row):
            col_name = header[idx] if idx < len(header) and header[idx] else f"列{idx}"
            fields.append(f"{col_name}: {val.strip()}")
        rows.append(" | ".join(fields))
    if not rows:
        return CleanResult("", ["CSV 无数据行"])
    return CleanResult("\n".join(rows))


def _clean_html(data: bytes) -> CleanResult:
    """BeautifulSoup 提取 body 可见文本，去除 script/style/nav/footer 等噪音标签。"""
    from bs4 import BeautifulSoup  # beautifulsoup4

    soup = BeautifulSoup(data, "lxml")
    # 移除噪音标签
    for tag in soup.find_all(["script", "style", "nav", "footer", "noscript", "iframe"]):
        tag.decompose()
    # 提取 body
    body = soup.find("body")
    target = body if body is not None else soup
    text = target.get_text(separator="\n", strip=True)
    # 合并空白行
    text = re.sub(r"\n{3,}", "\n\n", text)
    if not text.strip():
        return CleanResult("", ["HTML 内容为空"])
    return CleanResult(text)


def _clean_markdown(data: bytes) -> CleanResult:
    """Markdown ——规范化标题格式、去除多余空行。"""
    text = data.decode("utf-8", errors="replace")
    # 规范化带空格的标题：##  abc → ## abc
    text = re.sub(r"^(#{1,6})\s{2,}", r"\1 ", text, flags=re.MULTILINE)
    # 合并多余空行
    text = re.sub(r"\n{4,}", "\n\n\n", text)
    return CleanResult(text.strip())


def _clean_text(data: bytes) -> CleanResult:
    """纯文本 / JSON / YAML / TOML / 配置文件 ——段落合并、去连续空行。"""
    text = data.decode("utf-8", errors="replace")
    text = re.sub(r"\n{4,}", "\n\n\n", text)
    return CleanResult(text.strip())


def _clean_code(data: bytes) -> CleanResult:
    """代码文件——保留原样（高信号），仅去除连续多余空行。"""
    text = data.decode("utf-8", errors="replace")
    # 连续 4 个以上空行 → 3 个空行
    text = re.sub(r"\n{5,}", "\n\n\n\n", text)
    return CleanResult(text)


# ── 代码文件扩展名映射 ──────────────────────────────────────────────────────────

_CODE_EXTENSIONS: frozenset[str] = frozenset({
    "py", "ts", "tsx", "js", "jsx", "rs", "go", "java", "kt", "swift",
    "c", "cpp", "cxx", "h", "hpp", "cs", "rb", "php", "scala", "clj",
    "sql", "graphql", "proto", "vue", "svelte", "astro",
    "css", "scss", "less", "sh", "bash", "zsh", "ps1", "dockerfile",
    "xml", "rss", "svg",
})